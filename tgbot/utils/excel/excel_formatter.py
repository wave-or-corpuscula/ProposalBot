from openpyxl import *
from openpyxl.styles import *

from tgbot.utils.db_api.sqlite import DataBase
from tgbot.config import load_config
# from tgbot.utils.excel.excel_formatter import ExcelFormatter

class ExcelFormatter:

    def __init__(self, filename: str = "excel/Статистика.xlsx"):
        self.wb = Workbook()
        self.headers = ["Дата", "Тема", "Сообщение", "Ответ"]
        self.filename = filename
        self.start_row = 2
        self.start_col = 2

    def save_file(self):
        self.wb.save(self.filename)

    def fill_worksheet(self, sheetname: str, messages: list, border_style: str = "thin"):
        self.wb.create_sheet(sheetname)
        ws = self.wb[sheetname]

        thick_border = Border(left=Side(style=border_style),
                              right=Side(style=border_style),
                              top=Side(style=border_style),
                              bottom=Side(style=border_style))

        for i in range(len(self.headers)):
            cur_cell = ws.cell(row=self.start_row, column=self.start_col + i)
            cur_cell.value = self.headers[i]
            cur_cell.border = thick_border

        rows_amount = len(messages)
        cols_amount = len(self.headers)

        for i in range(rows_amount):
            for j in range(cols_amount):
                cur_cell = ws.cell(row=self.start_row + i + 1, column=self.start_col + j)
                cur_cell.value = messages[i][j]

                border = Border(right=Side(style=border_style), left=Side(style=border_style))
                if i == rows_amount - 1:
                    border.bottom = Side(style=border_style)
                cur_cell.border = border


    def fill_excel(self, messages: list, topics_messages: dict):
        self.wb.worksheets.clear()
        
        self.fill_worksheet("Все", messages)
        for topic_mess in topics_messages:
            for topic_name, mess in topic_mess.items():
                self.fill_worksheet(topic_name, mess)
        
        self.wb.remove_sheet(self.wb.active)
        self.save_file()

def test_main():
    from tgbot.utils.db_api.sqlite import DataBase
    from tgbot.config import load_config
    # from tgbot.utils.excel.excel_formatter import ExcelFormatter

    ex = ExcelFormatter()
    config = load_config(".env")
    db = DataBase(config)
    mess = db.get_all_messages_excel()
    topics_dict = db.get_topic_messages_excel()
    ex.fill_excel(mess, topics_dict)
